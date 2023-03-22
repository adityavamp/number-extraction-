
import argparse
import os
import platform
import sys
from pathlib import Path
from tkinter import filedialog as fd
from PIL import ImageEnhance
from PIL import Image, ImageOps 
import torch
import tkinter
import cv2
import numpy as np
from tkinter import *
from tkinter import ttk
from PIL import Image, ImageTk
from optparse import Option
from os import system
import tkinter
import cv2
import numpy as np
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter import *
from PIL import Image, ImageTk
import datetime
from openpyxl import load_workbook
import pandas as pd
from PIL import Image, ImageEnhance, ImageTk, ImageOps
from csv import writer
from keras.models import load_model
FILE = Path(__file__).resolve()
ROOT = FILE.parents[0]  # YOLOv5 root directory
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))  # add ROOT to PATH
ROOT = Path(os.path.relpath(ROOT, Path.cwd()))  # relative

from models.common import DetectMultiBackend
from utils.dataloaders import IMG_FORMATS, VID_FORMATS, LoadImages, LoadScreenshots, LoadStreams
from utils.general import (LOGGER, Profile, check_file, check_img_size, check_imshow, check_requirements, colorstr, cv2,
                           increment_path, non_max_suppression, print_args, scale_boxes, strip_optimizer, xyxy2xywh)
from utils.plots import Annotator, colors, save_one_box
from utils.torch_utils import select_device, smart_inference_mode




@smart_inference_mode()
def run(
        weights=ROOT / 'yolov5s.pt',  # model path or triton URL
        source=ROOT / '0',  # file/dir/URL/glob/screen/0(webcam)
        data=ROOT / 'data/coco128.yaml',  # dataset.yaml path
        imgsz=(640, 640),  # inference size (height, width)
        conf_thres=0.25,  # confidence threshold
        iou_thres=0.45,  # NMS IOU threshold
        max_det=1000,  # maximum detections per image
        device='',  # cuda device, i.e. 0 or 0,1,2,3 or cpu
        view_img=False,  # show results
        save_txt=False,  # save results to *.txt
        save_conf=False,  # save confidences in --save-txt labels
        save_crop=False,  # save cropped prediction boxes
        nosave=False,  # do not save images/videos
        classes=None,  # filter by class: --class 0, or --class 0 2 3
        agnostic_nms=False,  # class-agnostic NMS
        augment=False,  # augmented inference
        visualize=False,  # visualize features
        update=False,  # update all models
        project=ROOT / 'runs/detect',  # save results to project/name
        name='exp',  # save results to project/name
        exist_ok=False,  # existing project/name ok, do not increment
        line_thickness=3,  # bounding box thickness (pixels)
        hide_labels=False,  # hide labels
        hide_conf=False,  # hide confidences
        half=False,  # use FP16 half-precision inference
        dnn=False,  # use OpenCV DNN for ONNX inference
        vid_stride=1,  # video frame-rate stride
):
    source = str(source)
    save_img = not nosave and not source.endswith('.txt')  # save inference images
    is_file = Path(source).suffix[1:] in (IMG_FORMATS + VID_FORMATS)
    is_url = source.lower().startswith(('rtsp://', 'rtmp://', 'http://', 'https://'))
    webcam = source.isnumeric() or source.endswith('.txt') or (is_url and not is_file)
    screenshot = source.lower().startswith('screen')
    if is_url and is_file:
        source = check_file(source)  # download

    # Directories
    save_dir = increment_path(Path(project) / name, exist_ok=exist_ok)  # increment run
    (save_dir / 'labels' if save_txt else save_dir).mkdir(parents=True, exist_ok=True)  # make dir

    # Load model
    device = select_device(device)
    model = DetectMultiBackend(weights, device=device, dnn=dnn, data=data, fp16=half)
    stride, names, pt = model.stride, model.names, model.pt
    imgsz = check_img_size(imgsz, s=stride)  # check image size

    # Dataloader
    bs = 1  # batch_size
    if webcam:
        view_img = check_imshow(warn=True)
        dataset = LoadStreams(source, img_size=imgsz, stride=stride, auto=pt, vid_stride=vid_stride)
        bs = len(dataset)
    elif screenshot:
        dataset = LoadScreenshots(source, img_size=imgsz, stride=stride, auto=pt)
    else:
        dataset = LoadImages(source, img_size=imgsz, stride=stride, auto=pt, vid_stride=vid_stride)
    vid_path, vid_writer = [None] * bs, [None] * bs

    # Run inference
    print(partNumberSTR," ", partnamestr," ", CellNamestr, " " ,operationstr)
    model.warmup(imgsz=(1 if pt or model.triton else bs, 3, *imgsz))  # warmup
    seen, windows, dt = 0, [], (Profile(), Profile(), Profile())
    for path, im, im0s, vid_cap, s in dataset:
        with dt[0]:
            im = torch.from_numpy(im).to(model.device)
            im = im.half() if model.fp16 else im.float()  # uint8 to fp16/32
            im /= 255  # 0 - 255 to 0.0 - 1.0
            if len(im.shape) == 3:
                im = im[None]  # expand for batch dim

        # Inference
        with dt[1]:
            visualize = increment_path(save_dir / Path(path).stem, mkdir=True) if visualize else False
            pred = model(im, augment=augment, visualize=visualize)

        # NMS
        with dt[2]:
            pred = non_max_suppression(pred, conf_thres, iou_thres, classes, agnostic_nms, max_det=max_det)

        # Second-stage classifier (optional)
        # pred = utils.general.apply_classifier(pred, classifier_model, im, im0s)

        # Process predictions
        for i, det in enumerate(pred):  # per image
            seen += 1
            if webcam:  # batch_size >= 1
                p, im0, frame = path[i], im0s[i].copy(), dataset.count
                s += f'{i}: '
            else:
                p, im0, frame = path, im0s.copy(), getattr(dataset, 'frame', 0)

            p = Path(p)  # to Path
            save_path = str(save_dir / p.name)  # im.jpg
            txt_path = str(save_dir / 'labels' / p.stem) + ('' if dataset.mode == 'image' else f'_{frame}')  # im.txt
            s += '%gx%g ' % im.shape[2:]  # print string
            gn = torch.tensor(im0.shape)[[1, 0, 1, 0]]  # normalization gain whwh
            imc = im0.copy() if save_crop else im0  # for save_crop
            annotator = Annotator(im0, line_width=line_thickness, example=str(names))
            if len(det):
                # Rescale boxes from img_size to im0 size
                det[:, :4] = scale_boxes(im.shape[2:], det[:, :4], im0.shape).round()

                # Print results
                for c in det[:, 5].unique():
                    n = (det[:, 5] == c).sum()  # detections per class
                    s += f"{n} {names[int(c)]}{'s' * (n > 1)}, "  # add to string

                # Write results
                for *xyxy, conf, cls in reversed(det):
                    if save_txt:  # Write to file
                        xywh = (xyxy2xywh(torch.tensor(xyxy).view(1, 4)) / gn).view(-1).tolist()  # normalized xywh
                        line = (cls, *xywh, conf) if save_conf else (cls, *xywh)  # label format
                        with open(f'{txt_path}.txt', 'a') as f:
                            f.write(('%g ' * len(line)).rstrip() % line + '\n')

                    if save_img or save_crop or view_img:  # Add bbox to image
                        c = int(cls)  # integer class
                        label = None if hide_labels else (names[c] if hide_conf else f'{names[c]} {conf:.2f}')
                        # annotator.box_label(xyxy, label, color=colors(c, True))
                    
                    save_one_box(xyxy, imc,partNumberSTR, partnamestr, CellNamestr,operationstr, file=save_dir / 'crops' / names[c] / f'{p.stem}.jpg',BGR=True)

            # Stream results
            im0 = annotator.result()
            if view_img:
                if platform.system() == 'Linux' and p not in windows:
                    windows.append(p)
                    cv2.namedWindow(str(p), cv2.WINDOW_NORMAL | cv2.WINDOW_KEEPRATIO)  # allow window resize (Linux)
                    cv2.resizeWindow(str(p), im0.shape[1], im0.shape[0])
                cv2.imshow(str(p), im0)
                cv2.waitKey(1)  # 1 millisecond

            # Save results (image with detections)
            if save_img:
                if dataset.mode == 'image':
                    cv2.imwrite(save_path, im0)
                else:  # 'video' or 'stream'
                    if vid_path[i] != save_path:  # new video
                        vid_path[i] = save_path
                        if isinstance(vid_writer[i], cv2.VideoWriter):
                            vid_writer[i].release()  # release previous video writer
                        if vid_cap:  # video
                            fps = vid_cap.get(cv2.CAP_PROP_FPS)
                            w = int(vid_cap.get(cv2.CAP_PROP_FRAME_WIDTH))
                            h = int(vid_cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
                        else:  # stream
                            fps, w, h = 30, im0.shape[1], im0.shape[0]
                        save_path = str(Path(save_path).with_suffix('.mp4'))  # force *.mp4 suffix on results videos
                        vid_writer[i] = cv2.VideoWriter(save_path, cv2.VideoWriter_fourcc(*'mp4v'), fps, (w, h))
                    vid_writer[i].write(im0)

        # Print time (inference-only)
        LOGGER.info(f"{s}{'' if len(det) else '(no detections), '}{dt[1].dt * 1E3:.1f}ms")

    # Print results
    t = tuple(x.t / seen * 1E3 for x in dt)  # speeds per image
    LOGGER.info(f'Speed: %.1fms pre-process, %.1fms inference, %.1fms NMS per image at shape {(1, 3, *imgsz)}' % t)
    if save_txt or save_img:
        s = f"\n{len(list(save_dir.glob('labels/*.txt')))} labels saved to {save_dir / 'labels'}" if save_txt else ''
        LOGGER.info(f"Results saved to {colorstr('bold', save_dir)}{s}")
    if update:
        strip_optimizer(weights[0])  # update model (to fix SourceChangeWarning)


def parse_opt():
    parser = argparse.ArgumentParser()
    parser.add_argument('--weights', nargs='+', type=str, default=ROOT / 'best.pt', help='model path or triton URL')
    parser.add_argument('--source', type=str, default=ROOT / 'images', help='file/dir/URL/glob/screen/0(webcam)')
    parser.add_argument('--data', type=str, default=ROOT / 'ra.v1i.yolov5pytorch/data.yaml', help='(optional) dataset.yaml path')
    parser.add_argument('--imgsz', '--img', '--img-size', nargs='+', type=int, default=[1280], help='inference size h,w')
    parser.add_argument('--conf-thres', type=float, default=0.25, help='confidence threshold')
    parser.add_argument('--iou-thres', type=float, default=0.45, help='NMS IoU threshold')
    parser.add_argument('--max-det', type=int, default=1000, help='maximum detections per image')
    parser.add_argument('--device', default='', help='cuda device, i.e. 0 or 0,1,2,3 or cpu')
    parser.add_argument('--view-img', action='store_true', help='show results')
    parser.add_argument('--save-txt', action='store_true', help='save results to *.txt')
    parser.add_argument('--save-conf', action='store_true', help='save confidences in --save-txt labels')
    parser.add_argument('--save-crop', action='store_true', help='save cropped prediction boxes')
    parser.add_argument('--nosave', action='store_true', help='do not save images/videos')
    parser.add_argument('--classes', nargs='+', type=int, help='filter by class: --classes 0, or --classes 0 2 3')
    parser.add_argument('--agnostic-nms', action='store_true', help='class-agnostic NMS')
    parser.add_argument('--augment', action='store_true', help='augmented inference')
    parser.add_argument('--visualize', action='store_true', help='visualize features')
    parser.add_argument('--update', action='store_true', help='update all models')
    parser.add_argument('--project', default=ROOT / 'runs/detect', help='save results to project/name')
    parser.add_argument('--name', default='exp', help='save results to project/name')
    parser.add_argument('--exist-ok', action='store_true', help='existing project/name ok, do not increment')
    parser.add_argument('--line-thickness', default=3, type=int, help='bounding box thickness (pixels)')
    parser.add_argument('--hide-labels', default=False, action='store_true', help='hide labels')
    parser.add_argument('--hide-conf', default=False, action='store_true', help='hide confidences')
    parser.add_argument('--half', action='store_true', help='use FP16 half-precision inference')
    parser.add_argument('--dnn', action='store_true', help='use OpenCV DNN for ONNX inference')
    parser.add_argument('--vid-stride', type=int, default=1, help='video frame-rate stride')
    opt = parser.parse_args()
    opt.imgsz *= 2 if len(opt.imgsz) == 1 else 1  # expand
    print_args(vars(opt))
    return opt

def closeWindow():
    root.destroy()


#Main window
np.set_printoptions(suppress=True)

#loading new model


#getting current row count
def getRowCount():
    df = pd.read_excel("data_new.xlsx")
    rowCount = df['ID'].notna().sum()
    curent_row_count = rowCount + 1 
    return curent_row_count


#function to close the window
def closeWindow():
    root.destroy()


#Main window
root = Tk()
root.geometry("1640x950")
root.configure(bg="gray")
Label(root,text="Ra value Extracter", font=("times new roman", 24, "bold"),bg="white", fg="black").pack(pady=20)
f1 = LabelFrame(root,bg='gray')
f1.pack()
L1 = Label(f1,bg="gray")
L1.pack()
cap = cv2.VideoCapture(0)
#other part is at the bottom (click image button and close window button)


#Logo code



drop2 = ""
drop3 = ""
drop4 = ""
#<----------------------------------------- Drop Down Menu Section Starts ----------------------------------------->

#(1) Drop down menu for part number -------------------------------

#loading data from excel sheet
df = pd.read_excel("dropDownData/part_no_list.xlsx")
partNumberOptions = list(df["Part No"])
partNumberOptions = [str(i) for i in partNumberOptions]
partNumber = tkinter.StringVar()
partNumber.set("Select Part No.")

def _getPartNumber(curSelection):
    global partNumberSTR
    partNumberSTR = str(curSelection)
    drop2["state"] = "normal"

drop = tkinter.OptionMenu(root, partNumber, command=_getPartNumber, *partNumberOptions)
drop.place(x=75, y=630)

#--------------------------------------------------------------

#(2) drop down menu for machine number -------------------------------
machineNumberOptions = ['2040 SERIES', '2045 SERIES', '2055 SERIES', '2060 SERIES','TY','CF','HREY','FY 1410 series','FY 1350 series']
machineNumber = tkinter.StringVar()
machineNumber.set("Select Part Name.")
def _getMachineNumber(curSelection):
    global partnamestr
    partnamestr = str(curSelection)
    drop3["state"] = "normal"

drop2 = tkinter.OptionMenu(root, machineNumber, command=_getMachineNumber, *machineNumberOptions)
# drop2["state"] = "disabled"
drop2.place(x=280, y=630)

#--------------------------------------------------------------



#(3) Drop down menu for station number -------------------------------
stationNumberOptions = ['HDEY', 'GM', 'Chrysler', 'EY', 'TUNDRA', 'NISSAN', 'Mercedes', 'yoke shaft', 'Argentina', 'Jaguar','Hino']
stationNumber = tkinter.StringVar()
stationNumber.set("Select Cell Name.")
def _getStationNumber(curSelection):
    global CellNamestr
    CellNamestr = str(curSelection)
    drop4["state"] = "normal"

drop3 = tkinter.OptionMenu(root, stationNumber, command=_getStationNumber, *stationNumberOptions)
drop3.place(x=1000, y=630)
#--------------------------------------------------------------

#(4) Drop down menu for shift type -------------------------------
shiftTypeOptions = ['Grinding','Boring']
shiftType = tkinter.StringVar()
shiftType.set("Select Operation")
def _getShiftType(curSelection):
    global operationstr
    operationstr = curSelection 

drop4 = tkinter.OptionMenu(root, shiftType, command=_getShiftType, *shiftTypeOptions)
drop4.place(x=1170, y=630)
#--------------------------------------------------------------

#<----------------------------------------- Drop Down Menu Section Ends ----------------------------------------->


#function to click image and save it with datetime
def main(opt):
    check_requirements(exclude=('tensorboard', 'thop'))
    run(**vars(opt))

def clickImage():
    image = Image.fromarray(img1)
    image.save("saveimage/" + str(1) + ".jpg")
    opt = parse_opt()
    main(opt)


#function to reset dropdown data
def do_reset():
    partNumber.set("Select Part No.")
    machineNumber.set("Select Part Name.")
    stationNumber.set("Select Cell Name.")
    shiftType.set("Select Operation ")

def do_disable():
   drop2["state"] = "disable"
   drop3["state"] = "disable"
   drop4["state"] = "disable"




#function to select the file path
img_filepath = tkinter.StringVar()

#function to select image from a folder
def select_file():

    #excel sheet to check whether custom image is selected or not
    wrkbk = load_workbook("restricted/selectedStatus.xlsx")
    sh = wrkbk.active
    sh.cell(row=1,column=1).value = 1
    wrkbk.save("restricted/selectedStatus.xlsx")

    filetypes = (
        ('jpg files', '*.jpg'),
        ('png files', '*.png'),
         ('All files', '*.*')
    )

    filename = fd.askopenfilename(
        title='Window',
        initialdir='/',
        filetypes=filetypes)
    
    global selectedImagePath

    #variable selectedImagePath contains the filepath of the selected image
    selectedImagePath = filename
    print("\n\nselect file Path: " + selectedImagePath + "\n\n")

    system(f"python preview.py \"{selectedImagePath}\"")
    return 0



Button(root,text="Select Image",height=2, width=15, font=("times new roman", 20, "bold"),bg="white", fg="blue", command = lambda:[clickImage(), do_reset()]).pack(pady=80)

while True:
    img = cap.read()[1]
    img = cv2.flip(img,1)
    img1 = cv2.cvtColor(img,cv2.COLOR_BGR2RGB)
    img = ImageTk.PhotoImage(Image.fromarray(img1))
    L1['image'] = img
    root.update()




